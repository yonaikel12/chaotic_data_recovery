from __future__ import annotations

import argparse
import csv
import html
import io
import json
import logging
import re
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import numpy as np
import pandas as pd
from charset_normalizer import from_bytes

try:
    import clevercsv
except Exception:
    clevercsv = None

try:
    import dateparser
except Exception:
    dateparser = None

try:
    import ftfy
except Exception:
    ftfy = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from unidecode import unidecode
except Exception:
    unidecode = None


TEXT_EXTENSIONS = {".csv", ".txt", ".tsv"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


@dataclass(slots=True)
class RecoveryConfig:
    header_row: int | None = None
    footers_to_skip: int = 0
    preferred_sheet: str | None = None
    forced_encoding: str | None = None
    forced_delimiter: str | None = None
    candidate_delimiters: tuple[str, ...] = (",", ";", "\t", "|", ":")
    decimal_separators: list[str] = field(default_factory=lambda: [",", "."])
    thousands_separators: list[str] = field(default_factory=lambda: [".", ",", " ", "'", "\u00a0"])
    currency_symbols: list[str] = field(default_factory=lambda: ["$", "€", "£", "USD", "EUR", "COP", "MXN", "S/", "ARS"])
    languages: list[str] = field(default_factory=lambda: ["es", "en", "pt", "fr", "de", "it"])
    date_order: str = "DMY"
    subtotal_keywords: tuple[str, ...] = (
        "subtotal",
        "sub-total",
        "total",
        "gran total",
        "sumatoria",
        "acumulado",
        "neto",
    )
    normalize_column_names: bool = True
    empty_row_threshold: float = 0.95
    max_header_scan_rows: int = 15
    numeric_inference_threshold: float = 0.6
    date_inference_threshold: float = 0.6
    treat_parentheses_as_negative: bool = True
    default_encoding: str = "utf-8"
    log_path: str | None = None


@dataclass(slots=True)
class LibraryTrigger:
    library: str
    reason: str
    details: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class GarbageRowLog:
    source_row_number: int
    reason: str
    preview: str


@dataclass(slots=True)
class ColumnHealth:
    column_name: str
    semantic_type: str
    non_null_input: int
    successful_standardizations: int
    success_rate: float
    notes: str = ""


@dataclass(slots=True)
class RecoveryIssue:
    step: str
    severity: str
    message: str


@dataclass(slots=True)
class RecoveryReport:
    source_path: str
    source_type: str = "unknown"
    encoding: str | None = None
    delimiter: str | None = None
    selected_sheet: str | None = None
    rows_in: int = 0
    rows_out: int = 0
    columns_out: int = 0
    library_triggers: list[LibraryTrigger] = field(default_factory=list)
    garbage_rows: list[GarbageRowLog] = field(default_factory=list)
    column_health: list[ColumnHealth] = field(default_factory=list)
    issues: list[RecoveryIssue] = field(default_factory=list)

    def register_library_trigger(self, library: str, reason: str, **details: Any) -> None:
        signature = (library, reason, tuple(sorted(details.items())))
        existing = {
            (item.library, item.reason, tuple(sorted(item.details.items())))
            for item in self.library_triggers
        }
        if signature not in existing:
            self.library_triggers.append(LibraryTrigger(library=library, reason=reason, details=details))

    def log_garbage_row(self, source_row_number: int, reason: str, preview: str) -> None:
        self.garbage_rows.append(
            GarbageRowLog(
                source_row_number=source_row_number,
                reason=reason,
                preview=preview[:160],
            )
        )

    def log_issue(self, step: str, severity: str, message: str) -> None:
        self.issues.append(RecoveryIssue(step=step, severity=severity, message=message))

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["library_triggers"] = [asdict(item) for item in self.library_triggers]
        payload["garbage_rows"] = [asdict(item) for item in self.garbage_rows]
        payload["column_health"] = [asdict(item) for item in self.column_health]
        payload["issues"] = [asdict(item) for item in self.issues]
        return payload

    def library_usage_frame(self) -> pd.DataFrame:
        return pd.DataFrame([asdict(item) for item in self.library_triggers])

    def column_health_frame(self) -> pd.DataFrame:
        return pd.DataFrame([asdict(item) for item in self.column_health])

    def garbage_rows_frame(self) -> pd.DataFrame:
        return pd.DataFrame([asdict(item) for item in self.garbage_rows])

    def issues_frame(self) -> pd.DataFrame:
        return pd.DataFrame([asdict(item) for item in self.issues])

    def summary(self) -> dict[str, Any]:
        return {
            "source_path": self.source_path,
            "source_type": self.source_type,
            "encoding": self.encoding,
            "delimiter": self.delimiter,
            "selected_sheet": self.selected_sheet,
            "rows_in": self.rows_in,
            "rows_out": self.rows_out,
            "columns_out": self.columns_out,
            "libraries_fired": len(self.library_triggers),
            "garbage_rows_logged": len(self.garbage_rows),
            "issues_logged": len(self.issues),
        }

    def write_json(self, output_path: str | Path) -> None:
        Path(output_path).write_text(json.dumps(self.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


@dataclass(slots=True)
class RecoveryResult:
    dataframe: pd.DataFrame
    report: RecoveryReport


@dataclass(slots=True)
class SourceInspection:
    source_type: str
    encoding: str | None = None
    delimiter: str | None = None
    quotechar: str | None = None
    escapechar: str | None = None
    sheet_name: str | None = None


class DataRecoveryEngine:
    def __init__(self, config: RecoveryConfig | None = None, logger: logging.Logger | None = None) -> None:
        self.config = config or RecoveryConfig()
        self.logger = logger or self._build_logger()

    def run(self, source_path: str | Path) -> RecoveryResult:
        path = Path(source_path)
        report = RecoveryReport(source_path=str(path))
        inspection = self._safe_step(
            report,
            "inspect_source",
            lambda: self._inspect_source(path, report),
            fallback=SourceInspection(source_type=path.suffix.lower().lstrip(".") or "unknown"),
        )
        report.source_type = inspection.source_type
        report.encoding = inspection.encoding
        report.delimiter = inspection.delimiter
        report.selected_sheet = inspection.sheet_name

        raw_frame = self._safe_step(
            report,
            "load_raw_frame",
            lambda: self._load_raw_frame(path, inspection, report),
            fallback=pd.DataFrame(),
        )
        report.rows_in = int(raw_frame.shape[0])

        structured_frame = self._safe_step(
            report,
            "repair_structure",
            lambda: self._repair_structure(raw_frame, report),
            fallback=raw_frame.copy(),
        )

        cleaned_frame = self._safe_step(
            report,
            "clean_semantics",
            lambda: self._clean_semantics(structured_frame, report),
            fallback=structured_frame.copy(),
        )
        cleaned_frame = cleaned_frame.reset_index(drop=True)
        report.rows_out = int(cleaned_frame.shape[0])
        report.columns_out = int(cleaned_frame.shape[1])

        if self.config.log_path:
            self._safe_step(
                report,
                "write_report",
                lambda: report.write_json(self.config.log_path),
                fallback=None,
            )
        return RecoveryResult(dataframe=cleaned_frame, report=report)

    def _build_logger(self) -> logging.Logger:
        logger = logging.getLogger("chaotic_data_recovery")
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            logger.addHandler(logging.NullHandler())
        if self.config.log_path:
            file_handler = logging.FileHandler(self.config.log_path, encoding="utf-8")
            file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
            logger.addHandler(file_handler)
        return logger

    def _safe_step(self, report: RecoveryReport, step_name: str, func: Callable[[], Any], fallback: Any) -> Any:
        try:
            return func()
        except Exception as exc:
            self.logger.exception("Step %s failed", step_name)
            report.log_issue(step=step_name, severity="warning", message=str(exc))
            return fallback() if callable(fallback) else fallback

    def _inspect_source(self, path: Path, report: RecoveryReport) -> SourceInspection:
        suffix = path.suffix.lower()
        if suffix in TEXT_EXTENSIONS:
            raw_bytes = path.read_bytes()
            encoding = self._detect_encoding(raw_bytes, report)
            text = raw_bytes.decode(encoding or self.config.default_encoding, errors="replace")
            delimiter, quotechar, escapechar = self._detect_dialect(text, suffix, report)
            return SourceInspection(
                source_type="text",
                encoding=encoding,
                delimiter=delimiter,
                quotechar=quotechar,
                escapechar=escapechar,
            )
        if suffix in EXCEL_EXTENSIONS:
            sheet_name = self.config.preferred_sheet
            return SourceInspection(source_type="excel", sheet_name=sheet_name)
        return SourceInspection(source_type=suffix.lstrip(".") or "unknown")

    def _detect_encoding(self, raw_bytes: bytes, report: RecoveryReport) -> str:
        if self.config.forced_encoding:
            report.register_library_trigger(
                "config_override",
                "Used analyst-provided encoding override instead of byte-level inference.",
                encoding=self.config.forced_encoding,
            )
            return self.config.forced_encoding
        detection = from_bytes(raw_bytes[:50000]).best()
        encoding = detection.encoding if detection and detection.encoding else self.config.default_encoding
        report.register_library_trigger(
            "charset-normalizer",
            "Byte-level encoding inspection selected an input encoding.",
            encoding=encoding,
        )
        return encoding

    def _detect_dialect(self, text: str, suffix: str, report: RecoveryReport) -> tuple[str, str | None, str | None]:
        if self.config.forced_delimiter:
            report.register_library_trigger(
                "config_override",
                "Used analyst-provided delimiter override instead of dialect detection.",
                delimiter=self.config.forced_delimiter,
            )
            return self.config.forced_delimiter, '"', None
        if clevercsv is not None:
            try:
                sample = text[:20000]
                dialect = clevercsv.Sniffer().sniff(sample)
                delimiter = getattr(dialect, "delimiter", None) or ("\t" if suffix == ".tsv" else ",")
                quotechar = getattr(dialect, "quotechar", None)
                escapechar = getattr(dialect, "escapechar", None)
                report.register_library_trigger(
                    "clevercsv",
                    "Detected CSV dialect from inconsistent row patterns and delimiters.",
                    delimiter=delimiter,
                    quotechar=quotechar or "",
                    escapechar=escapechar or "",
                )
                return delimiter, quotechar, escapechar
            except Exception as exc:
                report.log_issue("detect_dialect", "warning", f"clevercsv fallback: {exc}")
        delimiter = self._heuristic_delimiter(text, suffix)
        report.register_library_trigger(
            "builtin_csv_fallback",
            "Applied heuristic delimiter detection because specialized dialect sniffing did not converge.",
            delimiter=delimiter,
        )
        return delimiter, '"', None

    def _heuristic_delimiter(self, text: str, suffix: str) -> str:
        if suffix == ".tsv":
            return "\t"
        sample_lines = [line for line in text.splitlines()[:25] if line.strip()]
        if not sample_lines:
            return ","
        scores: dict[str, int] = {}
        for delimiter in self.config.candidate_delimiters:
            counts = [line.count(delimiter) for line in sample_lines]
            if max(counts, default=0) == 0:
                continue
            scores[delimiter] = int(np.mean(counts) * 100 - np.std(counts) * 10)
        return max(scores, key=scores.get) if scores else ","

    def _load_raw_frame(self, path: Path, inspection: SourceInspection, report: RecoveryReport) -> pd.DataFrame:
        if inspection.source_type == "text":
            return self._load_text_frame(path, inspection, report)
        if inspection.source_type == "excel":
            return self._load_excel_frame(path, inspection, report)
        report.log_issue("load_raw_frame", "warning", f"Unsupported extension for fail-safe ingestion: {path.suffix}")
        return pd.DataFrame()

    def _load_text_frame(self, path: Path, inspection: SourceInspection, report: RecoveryReport) -> pd.DataFrame:
        encoding = inspection.encoding or self.config.default_encoding
        text = path.read_text(encoding=encoding, errors="replace")
        rows: list[list[Any]] = []
        if clevercsv is not None:
            try:
                if inspection.quotechar is not None:
                    reader = clevercsv.reader(
                        io.StringIO(text),
                        delimiter=inspection.delimiter or ",",
                        quotechar=inspection.quotechar,
                        escapechar=inspection.escapechar,
                    )
                else:
                    reader = clevercsv.reader(io.StringIO(text), delimiter=inspection.delimiter or ",")
                rows = [list(row) for row in reader]
            except Exception as exc:
                report.log_issue("load_text_frame", "warning", f"clevercsv reader fallback: {exc}")
        if not rows:
            reader = csv.reader(io.StringIO(text), delimiter=inspection.delimiter or ",")
            rows = [list(row) for row in reader]
        return self._rows_to_frame(rows)

    def _load_excel_frame(self, path: Path, inspection: SourceInspection, report: RecoveryReport) -> pd.DataFrame:
        if load_workbook is None:
            raise ImportError("openpyxl is required to read Excel sources.")
        workbook = load_workbook(filename=path, data_only=True)
        if inspection.sheet_name and inspection.sheet_name in workbook.sheetnames:
            worksheet = workbook[inspection.sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        report.selected_sheet = worksheet.title

        max_row = worksheet.max_row or 0
        max_column = worksheet.max_column or 0
        rows = [
            [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_column + 1)]
            for row_idx in range(1, max_row + 1)
        ]

        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            report.register_library_trigger(
                "openpyxl",
                "Expanded merged Excel cells by propagating anchor values across each merged range.",
                merged_ranges=len(merged_ranges),
                sheet=worksheet.title,
            )
        for merged_range in merged_ranges:
            anchor_value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    rows[row_idx - 1][col_idx - 1] = anchor_value
        return self._rows_to_frame(rows)

    def _rows_to_frame(self, rows: list[list[Any]]) -> pd.DataFrame:
        if not rows:
            return pd.DataFrame()
        width = max((len(row) for row in rows), default=0)
        padded = [row + [None] * (width - len(row)) for row in rows]
        frame = pd.DataFrame(padded, dtype="object")
        frame = frame.apply(
            lambda column: column.map(lambda value: np.nan if isinstance(value, str) and not value.strip() else value)
        )
        return frame

    def _repair_structure(self, frame: pd.DataFrame, report: RecoveryReport) -> pd.DataFrame:
        working = frame.copy()
        if working.empty:
            return working

        working = working.dropna(how="all").dropna(axis=1, how="all")
        if working.empty:
            return working

        if self.config.footers_to_skip > 0 and len(working) > self.config.footers_to_skip:
            footer_slice = working.tail(self.config.footers_to_skip)
            for source_idx, row in footer_slice.iterrows():
                report.log_garbage_row(int(source_idx) + 1, "configured_footer_skip", self._row_preview(row.tolist()))
            working = working.iloc[:-self.config.footers_to_skip]

        header_position = self.config.header_row
        if header_position is None:
            header_position = self._detect_header_position(working)
        header_position = max(0, min(header_position, len(working) - 1))

        preamble = working.iloc[:header_position]
        for source_idx, row in preamble.iterrows():
            report.log_garbage_row(int(source_idx) + 1, "preamble_or_title", self._row_preview(row.tolist()))

        header_values = self._make_unique_headers(working.iloc[header_position].tolist())
        body = working.iloc[header_position + 1 :].copy()
        body.columns = header_values

        candidates, discarded = self._extract_primary_table(body)
        for source_row_number, reason, preview in discarded:
            report.log_garbage_row(source_row_number, reason, preview)
        return candidates

    def _detect_header_position(self, frame: pd.DataFrame) -> int:
        scan_limit = min(len(frame), self.config.max_header_scan_rows)
        best_position = 0
        best_score = float("-inf")
        for position in range(scan_limit):
            row = frame.iloc[position].tolist()
            score = self._header_score(row)
            if score > best_score:
                best_score = score
                best_position = position
        return best_position

    def _header_score(self, row: list[Any]) -> float:
        values = [self._clean_text(cell) for cell in row if self._is_non_empty(cell)]
        if not values:
            return float("-inf")
        unique_ratio = len(set(values)) / max(len(values), 1)
        alpha_cells = sum(any(character.isalpha() for character in value) for value in values)
        numeric_cells = sum(self._looks_numeric(value) for value in values)
        long_note_penalty = sum(len(value) > 40 for value in values)
        return len(values) * 2.0 + unique_ratio * 3.0 + alpha_cells * 1.5 - numeric_cells * 1.2 - long_note_penalty * 2.0

    def _extract_primary_table(self, body: pd.DataFrame) -> tuple[pd.DataFrame, list[tuple[int, str, str]]]:
        if body.empty:
            return body, []

        kept_rows: list[dict[str, Any]] = []
        discarded: list[tuple[int, str, str]] = []
        rows_meta: list[tuple[bool, int, str, dict[str, Any], str]] = []

        for source_idx, row in body.iterrows():
            row_dict = row.to_dict()
            is_garbage, reason = self._classify_body_row(row.tolist())
            preview = self._row_preview(row.tolist())
            rows_meta.append((is_garbage, int(source_idx) + 1, reason, row_dict, preview))

        segments: list[tuple[int, int, float]] = []
        start: int | None = None
        for idx, item in enumerate(rows_meta):
            is_garbage = item[0]
            if not is_garbage and start is None:
                start = idx
            if is_garbage and start is not None:
                segments.append((start, idx - 1, self._segment_score(rows_meta[start:idx])))
                start = None
        if start is not None:
            segments.append((start, len(rows_meta) - 1, self._segment_score(rows_meta[start:])))

        if segments:
            best_start, best_end, _ = max(segments, key=lambda item: item[2])
        else:
            best_start, best_end = 0, len(rows_meta) - 1

        for idx, (is_garbage, source_row_number, reason, row_dict, preview) in enumerate(rows_meta):
            if best_start <= idx <= best_end and not is_garbage:
                kept_rows.append(row_dict)
            else:
                discard_reason = reason if is_garbage else "outside_primary_table_segment"
                discarded.append((source_row_number, discard_reason, preview))

        result = pd.DataFrame(kept_rows)
        result = result.dropna(how="all")
        return result.reset_index(drop=True), discarded

    def _segment_score(self, rows_meta: list[tuple[bool, int, str, dict[str, Any], str]]) -> float:
        if not rows_meta:
            return float("-inf")
        density = np.mean([sum(self._is_non_empty(value) for value in row_dict.values()) for _, _, _, row_dict, _ in rows_meta])
        return float(len(rows_meta) * density)

    def _classify_body_row(self, row: list[Any]) -> tuple[bool, str]:
        non_empty_values = [self._clean_text(value) for value in row if self._is_non_empty(value)]
        if not non_empty_values:
            return True, "empty_row"
        joined = " ".join(non_empty_values).lower()
        if any(keyword in joined for keyword in self.config.subtotal_keywords):
            return True, "subtotal_or_total_row"
        if len(non_empty_values) == 1 and len(non_empty_values[0]) > 24:
            return True, "decorative_or_note_row"
        if all(re.fullmatch(r"[-=*_/\\.\s]+", value or "") for value in non_empty_values):
            return True, "separator_row"
        row_width = len(non_empty_values) / max(len(row), 1)
        if row_width <= (1.0 - self.config.empty_row_threshold):
            return True, "sparse_noise_row"
        return False, "data_row"

    def _clean_semantics(self, frame: pd.DataFrame, report: RecoveryReport) -> pd.DataFrame:
        cleaned = frame.copy()
        if cleaned.empty:
            return cleaned

        renamed_columns = self._normalize_columns(list(cleaned.columns), report)
        cleaned.columns = renamed_columns

        for column in cleaned.columns:
            column_series = cleaned[column]
            non_null_count = int(column_series.notna().sum())
            if non_null_count == 0:
                report.column_health.append(
                    ColumnHealth(
                        column_name=column,
                        semantic_type="empty",
                        non_null_input=0,
                        successful_standardizations=0,
                        success_rate=1.0,
                        notes="Column was empty after structural repair.",
                    )
                )
                continue

            numeric_converted, numeric_success = self._convert_series_to_numeric(column_series, report)
            date_converted, date_success = self._convert_series_to_datetime(column_series, report)

            numeric_ratio = numeric_success / non_null_count
            date_ratio = date_success / non_null_count

            if numeric_ratio >= self.config.numeric_inference_threshold and numeric_ratio > date_ratio:
                cleaned[column] = numeric_converted
                report.column_health.append(
                    ColumnHealth(
                        column_name=column,
                        semantic_type="numeric",
                        non_null_input=non_null_count,
                        successful_standardizations=numeric_success,
                        success_rate=round(numeric_ratio, 4),
                        notes=f"numeric_ratio={numeric_ratio:.2f}; date_ratio={date_ratio:.2f}",
                    )
                )
            elif date_ratio >= self.config.date_inference_threshold:
                cleaned[column] = date_converted
                report.column_health.append(
                    ColumnHealth(
                        column_name=column,
                        semantic_type="datetime",
                        non_null_input=non_null_count,
                        successful_standardizations=date_success,
                        success_rate=round(date_ratio, 4),
                        notes=f"date_ratio={date_ratio:.2f}; numeric_ratio={numeric_ratio:.2f}",
                    )
                )
            else:
                text_converted, text_success = self._convert_series_to_text(column_series, report)
                cleaned[column] = text_converted
                report.column_health.append(
                    ColumnHealth(
                        column_name=column,
                        semantic_type="text",
                        non_null_input=non_null_count,
                        successful_standardizations=text_success,
                        success_rate=round(text_success / non_null_count, 4),
                        notes=f"numeric_ratio={numeric_ratio:.2f}; date_ratio={date_ratio:.2f}",
                    )
                )
        return cleaned

    def _normalize_columns(self, columns: list[Any], report: RecoveryReport) -> list[str]:
        normalized: list[str] = []
        for index, column in enumerate(columns, start=1):
            text = self._clean_text(column) if self.config.normalize_column_names else str(column)
            if self.config.normalize_column_names and unidecode is not None:
                transliterated = unidecode(text)
                if transliterated != text:
                    report.register_library_trigger(
                        "Unidecode",
                        "Normalized column labels to ASCII-friendly identifiers.",
                        column=str(column),
                    )
                text = transliterated
            text = re.sub(r"[^0-9a-zA-Z]+", "_", text).strip("_").lower()
            if not text:
                text = f"column_{index}"
            normalized.append(text)
        return self._make_unique_headers(normalized)

    def _make_unique_headers(self, values: list[Any]) -> list[str]:
        prepared: list[str] = []
        counts: Counter[str] = Counter()
        for index, value in enumerate(values, start=1):
            text = self._clean_text(value)
            text = re.sub(r"\s+", " ", text).strip()
            text = text if text else f"column_{index}"
            text = re.sub(r"[^0-9a-zA-Z_ ]+", "", text).strip() or f"column_{index}"
            candidate = text
            counts[candidate] += 1
            if counts[candidate] > 1:
                candidate = f"{candidate}_{counts[candidate]}"
            prepared.append(candidate)
        return prepared

    def _convert_series_to_numeric(self, series: pd.Series, report: RecoveryReport) -> tuple[pd.Series, int]:
        converted_values: list[Any] = []
        success_count = 0
        for value in series.tolist():
            parsed, success = self._parse_numeric(value)
            converted_values.append(parsed)
            success_count += int(success)
        converted = pd.Series(converted_values, index=series.index, dtype="float64")
        if success_count > 0:
            report.register_library_trigger(
                "numeric_normalizer",
                "Standardized mixed decimal, thousands, currency, and parenthesis-based numeric formats.",
                successful_values=success_count,
            )
        return converted, success_count

    def _convert_series_to_datetime(self, series: pd.Series, report: RecoveryReport) -> tuple[pd.Series, int]:
        converted_values: list[Any] = []
        success_count = 0
        for value in series.tolist():
            parsed, success = self._parse_date(value)
            converted_values.append(parsed)
            success_count += int(success)
        converted = pd.to_datetime(pd.Series(converted_values, index=series.index), errors="coerce")
        if success_count > 0 and dateparser is not None:
            report.register_library_trigger(
                "dateparser",
                "Parsed multilingual or ambiguous date strings into normalized timestamps.",
                successful_values=success_count,
                date_order=self.config.date_order,
            )
        return converted, success_count

    def _convert_series_to_text(self, series: pd.Series, report: RecoveryReport) -> tuple[pd.Series, int]:
        cleaned_values: list[Any] = []
        success_count = 0
        changed_count = 0
        for value in series.tolist():
            if pd.isna(value):
                cleaned_values.append(np.nan)
                continue
            original = re.sub(r"\s+", " ", str(value)).strip()
            try:
                cleaned_text = self._clean_text(value)
                cleaned_values.append(cleaned_text if cleaned_text else np.nan)
                success_count += 1
                changed_count += int(cleaned_text != original)
            except Exception:
                cleaned_values.append(np.nan)
        if changed_count > 0 and ftfy is not None:
            report.register_library_trigger(
                "ftfy",
                "Repaired mojibake, HTML entities, or broken punctuation in text fields.",
                repaired_values=changed_count,
            )
        return pd.Series(cleaned_values, index=series.index, dtype="object"), success_count

    def _parse_numeric(self, value: Any) -> tuple[float | None, bool]:
        if pd.isna(value):
            return np.nan, False
        if isinstance(value, (int, float, np.number)) and not isinstance(value, bool):
            return float(value), True

        text = self._clean_text(value)
        if not text or not re.search(r"\d", text):
            return np.nan, False
        if self._looks_date_like(text):
            return np.nan, False

        negative = False
        if self.config.treat_parentheses_as_negative and text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1]

        for symbol in self.config.currency_symbols:
            text = text.replace(symbol, "")
        text = text.replace("%", "")
        text = text.replace("\u2212", "-")
        text = re.sub(r"\s+", "", text)

        decimal_separator = self._infer_decimal_separator(text)
        thousands_separators = [separator for separator in self.config.thousands_separators if separator != decimal_separator]

        if decimal_separator and decimal_separator in text:
            integer_part, fractional_part = text.rsplit(decimal_separator, 1)
            for separator in thousands_separators:
                integer_part = integer_part.replace(separator, "")
                fractional_part = fractional_part.replace(separator, "")
            text = integer_part + "." + re.sub(r"[^0-9]", "", fractional_part)
        else:
            for separator in self.config.thousands_separators:
                text = text.replace(separator, "")

        text = re.sub(r"[^0-9\-.]", "", text)
        text = re.sub(r"(?<!^)-", "", text)
        if negative and not text.startswith("-"):
            text = "-" + text
        if text in {"", "-", ".", "-."}:
            return np.nan, False

        try:
            return float(text), True
        except ValueError:
            return np.nan, False

    def _infer_decimal_separator(self, text: str) -> str | None:
        candidates: list[tuple[str, int]] = []
        for separator in self.config.decimal_separators:
            if separator not in text:
                continue
            last_position = text.rfind(separator)
            fractional = text[last_position + 1 :]
            if fractional.isdigit() and 1 <= len(fractional) <= 4:
                candidates.append((separator, last_position))
        if candidates:
            return max(candidates, key=lambda item: item[1])[0]
        return None

    def _parse_date(self, value: Any) -> tuple[pd.Timestamp | pd.NaT, bool]:
        if pd.isna(value):
            return pd.NaT, False
        if isinstance(value, pd.Timestamp):
            return value.tz_localize(None) if value.tzinfo else value, True
        if isinstance(value, datetime):
            return pd.Timestamp(value.replace(tzinfo=None)), True
        if isinstance(value, date):
            return pd.Timestamp(value), True
        if dateparser is None:
            return pd.NaT, False

        text = self._clean_text(value)
        if not text or not re.search(r"\d", text):
            return pd.NaT, False
        if any(symbol in text for symbol in self.config.currency_symbols):
            return pd.NaT, False

        parsed = dateparser.parse(
            text,
            languages=self.config.languages,
            settings={
                "DATE_ORDER": self.config.date_order,
                "PREFER_DAY_OF_MONTH": "first",
                "RETURN_AS_TIMEZONE_AWARE": False,
            },
        )
        if parsed is None:
            return pd.NaT, False
        return pd.Timestamp(parsed), True

    def _clean_text(self, value: Any) -> str:
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return ""
        if isinstance(value, str):
            text = html.unescape(value)
            if ftfy is not None:
                text = ftfy.fix_text(text)
            return re.sub(r"\s+", " ", text).strip()
        return re.sub(r"\s+", " ", str(value)).strip()

    def _looks_numeric(self, value: str) -> bool:
        compact = value.replace(" ", "")
        return bool(re.fullmatch(r"[\(\-]?[\d\.,%]+\)?", compact))

    def _looks_date_like(self, value: str) -> bool:
        lowered = value.lower()
        month_tokens = {
            "ene", "enero", "feb", "febrero", "mar", "marzo", "abr", "abril", "may", "mayo",
            "jun", "junio", "jul", "julio", "ago", "agosto", "sep", "sept", "septiembre",
            "oct", "octubre", "nov", "noviembre", "dic", "diciembre", "jan", "february",
            "march", "april", "june", "july", "august", "september", "october", "november", "december",
        }
        if any(token in lowered for token in month_tokens):
            return True
        return bool(re.fullmatch(r"\d{1,4}[/-]\d{1,2}[/-]\d{1,4}", lowered))

    def _is_non_empty(self, value: Any) -> bool:
        return not pd.isna(value) and str(value).strip() != ""

    def _row_preview(self, row: list[Any]) -> str:
        values = [self._clean_text(value) for value in row if self._is_non_empty(value)]
        return " | ".join(values[:5])


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fail-safe ingestion and cleaning for chaotic business files.")
    parser.add_argument("source_path", help="Path to the chaotic file to process.")
    parser.add_argument("--header-row", type=int, default=None, help="Zero-based row index containing headers.")
    parser.add_argument("--footers-to-skip", type=int, default=0, help="Number of footer rows to ignore.")
    parser.add_argument("--sheet", dest="preferred_sheet", default=None, help="Preferred Excel sheet name.")
    parser.add_argument("--output-csv", default=None, help="Optional path to save the cleaned DataFrame.")
    parser.add_argument("--output-report", default=None, help="Optional path to save the quality report JSON.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)
    config = RecoveryConfig(
        header_row=args.header_row,
        footers_to_skip=args.footers_to_skip,
        preferred_sheet=args.preferred_sheet,
        log_path=args.output_report,
    )
    engine = DataRecoveryEngine(config=config)
    result = engine.run(args.source_path)
    if args.output_csv:
        result.dataframe.to_csv(args.output_csv, index=False)
    if args.output_report:
        result.report.write_json(args.output_report)
    print(json.dumps(result.report.summary(), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
