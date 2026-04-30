from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook

from chaotic_data_recovery import DataRecoveryEngine, RecoveryConfig


def _build_frankenstein_excel(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "caos_regional"

    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = "REPORTE COMERCIAL REGIONAL"
    worksheet.merge_cells("A2:B2")
    worksheet["A2"] = "Corte marzo 2026"
    worksheet["D2"] = "Documento interno"
    worksheet["A3"] = "-----------------------------------------"
    worksheet.append(["Cliente", "Fecha", "Monto", "Canal", "Notas"])
    worksheet.append(["JosÃ© Perez", "15 ene 2025", "€ 1.234,50", "Retail", "Pago en BogotÃ¡"])
    worksheet.append(["Marta Nunez", "03/07/2024", "USD 2,450.90", "Digital", "Cliente VIP"])
    worksheet.append(["Luis Gomez", "2024-11-05", "(3.000,00)", "Mayorista", "NC aplicada"])
    worksheet.append(["TOTAL", None, "$ 685,40", None, None])
    worksheet.append(["Nota legal: cifras preliminares sujetas a validacion", None, None, None, None])

    workbook.save(path)
    return path


def _build_cp1252_csv(path: Path, *, include_relative_date: bool) -> Path:
    relative_or_explicit_date = "ayer" if include_relative_date else "12/02/2026"
    rows = [
        "REPORTE EXPORTADO POR VENTAS - region norte;;;;",
        "No modificar manualmente | senal verde;;;;",
        "Cliente;Fecha;Monto;Moneda;Observacion",
        "Ana Peña;7 marzo 2025;€ 1.234,50;EUR;Cobro confirmado",
        "Carlos Ruiz;04/11/2024;USD 2,450.90;USD;Pago parcial",
        f"Elena Díaz;{relative_or_explicit_date};EUR 3 520,10;EUR;Registro tardío",
        "TOTAL;;;;",
        "Pie: generado desde CRM legado versión 2;;;;",
    ]
    path.write_text("\n".join(rows) + "\n", encoding="cp1252", errors="replace")
    return path


@pytest.fixture
def frankenstein_excel_path(tmp_path: Path) -> Path:
    return _build_frankenstein_excel(tmp_path / "frankenstein_business_report.xlsx")


@pytest.fixture
def cp1252_csv_factory(tmp_path: Path):
    def _factory(filename: str, *, include_relative_date: bool) -> Path:
        return _build_cp1252_csv(tmp_path / filename, include_relative_date=include_relative_date)

    return _factory


def test_excel_recovery_handles_merged_cells_mojibake_and_garbage_rows(frankenstein_excel_path: Path) -> None:
    engine = DataRecoveryEngine(RecoveryConfig(footers_to_skip=1))

    result = engine.run(frankenstein_excel_path)
    libraries = {item.library for item in result.report.library_triggers}
    garbage_reasons = {item.reason for item in result.report.garbage_rows}

    assert list(result.dataframe.columns) == ["cliente", "fecha", "monto", "canal", "notas"]
    assert result.report.selected_sheet == "caos_regional"
    assert result.report.rows_out == 3
    assert result.dataframe["cliente"].tolist() == ["José Perez", "Marta Nunez", "Luis Gomez"]
    assert result.dataframe["monto"].tolist() == pytest.approx([1234.5, 2450.9, -3000.0])
    assert result.dataframe["notas"].tolist()[0] == "Pago en Bogotá"
    assert result.dataframe["fecha"].dt.strftime("%Y-%m-%d").tolist() == ["2025-01-15", "2024-07-03", "2024-05-11"]
    assert {"openpyxl", "ftfy", "dateparser", "numeric_normalizer"}.issubset(libraries)
    assert {"preamble_or_title", "subtotal_or_total_row", "configured_footer_skip"}.issubset(garbage_reasons)


def test_csv_recovery_respects_encoding_and_delimiter_overrides(cp1252_csv_factory) -> None:
    csv_path = cp1252_csv_factory("frankenstein_sales_export_clean_dates.csv", include_relative_date=False)
    engine = DataRecoveryEngine(
        RecoveryConfig(
            header_row=2,
            footers_to_skip=1,
            forced_encoding="cp1252",
            forced_delimiter=";",
        )
    )

    result = engine.run(csv_path)
    triggers = [(item.library, item.reason) for item in result.report.library_triggers]

    assert result.report.encoding == "cp1252"
    assert result.report.delimiter == ";"
    assert result.dataframe["cliente"].tolist() == ["Ana Peña", "Carlos Ruiz", "Elena Díaz"]
    assert result.dataframe["fecha"].dt.strftime("%Y-%m-%d").tolist() == ["2025-03-07", "2024-11-04", "2026-02-12"]
    assert result.dataframe["monto"].tolist() == pytest.approx([1234.5, 2450.9, 3520.1])
    assert (
        "config_override",
        "Used analyst-provided encoding override instead of byte-level inference.",
    ) in triggers
    assert (
        "config_override",
        "Used analyst-provided delimiter override instead of dialect detection.",
    ) in triggers


def test_partial_date_parsing_is_recorded_without_stopping(cp1252_csv_factory) -> None:
    csv_path = cp1252_csv_factory("frankenstein_sales_export_partial_dates.csv", include_relative_date=True)
    engine = DataRecoveryEngine(
        RecoveryConfig(
            header_row=2,
            footers_to_skip=1,
            forced_encoding="cp1252",
        )
    )

    result = engine.run(csv_path)
    fecha_health = next(item for item in result.report.column_health if item.column_name == "fecha")

    assert result.report.rows_out == 3
    assert pd.isna(result.dataframe.loc[2, "fecha"])
    assert fecha_health.semantic_type == "datetime"
    assert fecha_health.successful_standardizations == 2
    assert fecha_health.success_rate == pytest.approx(0.6667, rel=1e-4)
    assert result.report.issues == []


def test_run_remains_fail_safe_when_semantic_cleaning_raises(tmp_path: Path) -> None:
    csv_path = tmp_path / "simple.csv"
    csv_path.write_text("Cliente,Monto\nAna,100\nLuis,250\n", encoding="utf-8")

    engine = DataRecoveryEngine(RecoveryConfig(header_row=0))

    with patch.object(DataRecoveryEngine, "_clean_semantics", side_effect=RuntimeError("synthetic semantic failure")):
        result = engine.run(csv_path)

    assert result.dataframe.to_dict(orient="records") == [
        {"Cliente": "Ana", "Monto": "100"},
        {"Cliente": "Luis", "Monto": "250"},
    ]
    assert any(issue.step == "clean_semantics" for issue in result.report.issues)
    assert any("synthetic semantic failure" in issue.message for issue in result.report.issues)